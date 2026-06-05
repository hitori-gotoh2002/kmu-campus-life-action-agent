"""ON국민 수강내역(수강신청확인서) 엑셀 파서 (결정론).

- `.xls`(xlrd) / `.xlsx`(openpyxl) 모두 지원, **여러 학기 파일 병합**.
- 헤더행을 스캔으로 탐지(고정행 X): '교과목코드'가 있는 행을 헤더로.
- 컬럼은 헤더 라벨로 매핑(병합셀로 인한 빈칸 무시).
- 교과목코드는 7자리 문자열로 보존(leading-zero·문자).
- 같은 코드가 여러 학기 파일에 나오면 `possible_retake`로 표시(성적 컬럼 없음 → 사용자 판단).
"""
from __future__ import annotations

import io

from graduation_center.v2.models_v2 import RawLine
from graduation_center.v2.text_norm import normalize_code

REQUIRED_COLS = ["교과목코드", "교과목명", "학점", "이수구분"]
_LABELS = {
    "교과목코드": "course_code", "분반": "section", "교과목명": "course_name",
    "이수구분": "area_raw", "학점": "credits", "담당교수": "professor", "비고": "note",
}


def _grid_from_xls(content: bytes) -> list[list[str]]:
    import xlrd
    book = xlrd.open_workbook(file_contents=content)
    sh = book.sheet_by_index(0)
    return [[_s(sh.cell_value(r, c)) for c in range(sh.ncols)] for r in range(sh.nrows)]


def _grid_from_xlsx(content: bytes) -> list[list[str]]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [[_s(v) for v in row] for row in ws.iter_rows(values_only=True)]


def _s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _grid(content: bytes, filename: str) -> list[list[str]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _grid_from_xlsx(content)
    if name.endswith(".xls"):
        return _grid_from_xls(content)
    # 매직바이트로 추정 (xlsx=zip PK, xls=OLE D0CF)
    if content[:2] == b"PK":
        return _grid_from_xlsx(content)
    return _grid_from_xls(content)


def _find_header(grid: list[list[str]]) -> int | None:
    for i, row in enumerate(grid):
        if any(c == "교과목코드" for c in row):
            return i
        # 실파일 변형: '학생수강신청내역출력'은 코드 컬럼명이 '교과목'(교과목명과 병존) — alias
        if "교과목" in row and "교과목명" in row:
            return i
    return None


def _header_term(grid: list[list[str]]) -> str:
    """헤더 위쪽에서 수강학기 라벨 추출 (예: '2023학년도 1학기')."""
    for row in grid[:6]:
        for j, c in enumerate(row):
            if c == "수강학기":
                for k in range(j + 1, len(row)):
                    if row[k]:
                        return row[k]
    return ""


def _alias_header(header: list[str]) -> list[str]:
    """실파일 헤더 변형 흡수 — ON국민 '학생수강신청내역출력'은 코드 컬럼명이 '교과목'
    (2026-06-05 실파일 확인). '교과목명'이 따로 있을 때만 '교과목'을 교과목코드로 간주."""
    if "교과목코드" not in header and "교과목" in header and "교과목명" in header:
        return ["교과목코드" if c == "교과목" else c for c in header]
    return header


def fail_fast_columns(content: bytes, filename: str) -> list[str]:
    """필수 컬럼 존재만 체크. 없는 컬럼 목록 반환(있으면 422)."""
    grid = _grid(content, filename)
    h = _find_header(grid)
    if h is None:
        return REQUIRED_COLS[:]
    header = _alias_header(grid[h])
    return [c for c in REQUIRED_COLS if c not in header]


def parse_file(content: bytes, filename: str) -> list[RawLine]:
    grid = _grid(content, filename)
    h = _find_header(grid)
    if h is None:
        return []
    header = _alias_header(grid[h])
    colmap = {}
    for idx, label in enumerate(header):
        if label in _LABELS:
            colmap[_LABELS[label]] = idx
    term_label = _header_term(grid) or (filename or "")
    lines: list[RawLine] = []
    for row in grid[h + 1:]:
        first = row[colmap.get("course_code", 0)] if row else ""
        if not first or first == "계":  # 합계행/빈행에서 종료
            if first == "계":
                break
            continue
        def g(key):
            i = colmap.get(key)
            return row[i] if i is not None and i < len(row) else ""
        try:
            credits = float(g("credits") or 0)
        except ValueError:
            credits = 0.0
        lines.append(RawLine(
            course_code=normalize_code(g("course_code")),
            course_name=g("course_name"),
            area_raw=g("area_raw"),
            credits=credits,
            section=g("section"),
            professor=g("professor"),
            note=g("note"),
            term_label=term_label,
        ))
    return lines


# 학사규정상 반복수강이 허용돼 매 학기 이수분이 각각 인정되는 과목(재수강 아님).
# 동일 코드가 여러 학기 나와도 '재수강 의심'으로 표시하지 않고 전부 포함한다.
REPEATABLE_COURSE_KEYWORDS = ("사제동행세미나",)


def _is_repeatable(name: str) -> bool:
    n = (name or "").replace(" ", "")
    return any(k.replace(" ", "") in n for k in REPEATABLE_COURSE_KEYWORDS)


def parse_many(files: list[tuple[bytes, str]]) -> tuple[list[RawLine], list[dict]]:
    """여러 학기 파일 → 병합 RawLine + possible_retakes(동일 코드 복수 학기).

    반복수강 허용 과목(사제동행세미나 등)은 재수강 후보에서 제외 — 매 이수분 인정.
    """
    all_lines: list[RawLine] = []
    for content, filename in files:
        all_lines.extend(parse_file(content, filename))
    seen: dict[str, list[str]] = {}
    repeatable_codes: set[str] = set()
    for ln in all_lines:
        if ln.course_code:
            seen.setdefault(ln.course_code, []).append(ln.term_label)
            if _is_repeatable(ln.course_name):
                repeatable_codes.add(ln.course_code)
    retakes = [{"course_code": code, "term_labels": terms}
               for code, terms in seen.items()
               if len(terms) > 1 and code not in repeatable_codes]
    return all_lines, retakes
